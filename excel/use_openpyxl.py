from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter

workbook = load_workbook('file/yangben.xlsx')
sheetnames = workbook.get_sheet_names()
print(sheetnames)

sheet = workbook.get_sheet_by_name(sheetnames[0])
print(sheet.cell(row=3,column=3).value)

sheet['A1']='use_openpyxl'

workbook.save('file/yangben_openpyxl.xlsx')
wb = Workbook()
ws = wb.active
ws['A1'] = 'use_openpyxl_2'
wb.save('file/yangben_openpyxl_2.xlsx')